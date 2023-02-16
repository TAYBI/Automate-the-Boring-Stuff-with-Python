import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

PRICE_UPDATE = {
    'Garlic': 3.07,
    'Celery': 1.19,
    'Lemon': 1.27
}

# Loop through the rows and update the prices.
for rowNum in range(2, sheet.max_row):
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in PRICE_UPDATE:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATE[produceName]

wb.save('updatedProduceSales.xlsx')
