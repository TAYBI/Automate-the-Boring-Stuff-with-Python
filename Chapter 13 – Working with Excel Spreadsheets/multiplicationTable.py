import sys
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
wb = openpyxl.Workbook()
sheet = wb.active

number = int(sys.argv[1])


for i in range(1, number):
    letter = get_column_letter(i)
    for j in range(1, number):
        # sheet[letter + str(i)] = i*j
        # print(letter + str(j) + f'={str(i*j)}\t', end="")
        sheet[letter + str(j)] = str(i*j)
    # print()

wb.save(f'multiplicationTable{number}.xlsx')
