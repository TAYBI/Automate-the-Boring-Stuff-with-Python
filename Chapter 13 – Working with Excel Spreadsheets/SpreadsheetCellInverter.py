import openpyxl
from openpyxl.utils import get_column_letter

print('openning workbook...')
wb = openpyxl.load_workbook('spreadsheet.xlsx')

sheet = wb['Sheet']
columns = []
rows = []

for column in range(1, sheet.max_column + 1):
    columns.append(get_column_letter(column))
for row in range(1, sheet.max_row + 1):
    rows.append(row)

DATA = {}

for column in columns:
    for row in rows:
        key = f'{column}{row}'
        value = sheet[f'{column}{row}'].value
        DATA[key] = value

UPDATEDDATA = {}

for i in range(1, len(rows)):
    for j in range(0, len(columns)):
        value = f'{columns[j]}{rows[i-1]}'
        key = f'{get_column_letter(i)}{j+1}'
        UPDATEDDATA[f'{get_column_letter(i)}{j+1}'] = sheet[value].value

# for key, value in DATA.items():
#     print(key, ':', value)
for key, value in UPDATEDDATA.items():
    print(key, ':', value)
